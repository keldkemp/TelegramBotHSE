"""
Класс в основном нужен для работы с сайтом HSE.
Сейчас здесь происходит основной движ по расписанию
"""
import requests
import pandas as pd
import os
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl.utils import range_boundaries
from Utils.utils import Utils
from models.Models import Pars
from openpyxl import load_workbook


class HSE:
    BASE_URL = 'https://perm.hse.ru'
    FILE_NAME = 'file_test.xls'
    NEW_FILE_NAME = 'file_test.xlsx'
    SHEET_NAME = '1 курс (СПО)'
    RU_MONTH_VALUES = {
        'января': '01',
        'февраля': '02',
        'марта': '03',
        'апреля': '04',
        'мая': '05',
        'июня': '06',
        'июля': '07',
        'августа': '08',
        'сентября': '09',
        'октября': 10,
        'ноября': 11,
        'декабря': 12,
    }

    def __merged_cells(self, sheet_name):
        wbook = load_workbook(filename=self.NEW_FILE_NAME)
        sheet = wbook[sheet_name]
        merge = [mcr.coord for mcr in sheet.merged_cells.ranges]
        for cell_group in merge:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
        wbook.save(self.NEW_FILE_NAME)
        wbook.close()

    def __get_days(self, t1: list) -> list:
        days = []
        i = 0
        for d in t1:
            if i < 2:
                i += 1
                days.append(d)
                continue
            if str(d) != 'nan':
                day = str(d)
            days.append(day)

        return days

    def __get_times(self, t2: list) -> list:
        times = []
        i = 0
        for d in t2:
            if i < 2:
                i += 1
                times.append(d)
                continue
            if str(d) != 'nan':
                time = str(d)
            times.append(time)

        return times

    def __get_pars(self, days: list, times: list, group1: list) -> list:
        pars = []
        flag = False
        i = 0
        for day, time, s1 in zip(days, times, group1):
            i += 1
            if flag:
                if str(s1) != 'nan':
                    pars.append(par1 + ";" + str(time) + ";" + str(s1))
                else:
                    pars.append(par1 + ";" + str(time) + ";" + str(s1))
                flag = False
                continue
            if str(s1).find('подразделения') != -1:
                continue
            if str(s1).find('учебного года') != -1:
                continue
            if i == 2:
                group_name1 = s1
                continue
            if str(s1).find('nan') != -1:
                continue
            if str(s1).find('nan') == -1:
                day = day[day.find('\n') + 1:]
                par1 = str(day) + ";" + str(s1) + ";" + group_name1
                flag = True
                continue

        return pars

    def __get_pars_data(self, pars: list) -> list:
        pars_data = []
        date_now = datetime.now()
        for par in pars:
            par = par.split(';')
            date = par[0].split()
            month = self.RU_MONTH_VALUES[date[1]]
            if len(str(date[0])) != 1:
                str_date = str(date_now.year) + '-' + str(month) + '-' + str(date[0])
            else:
                str_date = str(date_now.year) + '-' + str(month) + '-' + '0' + str(date[0])
            pars_data.append(Pars(str_date, par[1], par[2], par[3], par[4]))

        return pars_data

    def __parse_xls(self, file_name: str) -> list:
        # self.convert_xls_to_xlsx_site(file_name)
        Utils.convert_xls_to_xlsx(file_name, self.NEW_FILE_NAME)

        data_xlsx_sheets = pd.ExcelFile(self.NEW_FILE_NAME)
        sheets = data_xlsx_sheets.sheet_names
        pars_data = []

        for sheet in sheets:
            self.__merged_cells(sheet_name=sheet)
            data_xlsx = pd.read_excel(self.NEW_FILE_NAME, sheet, index_col=None)
            data_xls = pd.read_excel(file_name, sheet, index_col=None)

            # 0 - всегда дни, 1 - всегда время. Остальные в списке - это группы
            head_column = list(data_xlsx.columns)
            # ar_js = json.loads(data_xlsx.to_json())

            t1 = data_xls[head_column[0]].tolist()
            t2 = data_xls[head_column[1]].tolist()
            head_column.pop(0)
            head_column.pop(0)

            for col_name in head_column:
                group = data_xlsx[col_name].tolist()

                days = self.__get_days(t1)
                times = self.__get_times(t2)
                pars = self.__get_pars(days=days, times=times, group1=group)
                pars_data.append(self.__get_pars_data(pars))

        return pars_data

    def download_files(self, urls: list) -> list:
        pars_data = []
        for url in urls:
            f = open(self.FILE_NAME, 'wb')
            ufr = requests.get(url)
            f.write(ufr.content)
            f.close()
            pars_data.append(self.__parse_xls(self.FILE_NAME))
        os.remove(self.FILE_NAME)
        os.remove(self.NEW_FILE_NAME)
        return pars_data

    def get_urls(self, html) -> list:
        soup = BeautifulSoup(html.text, 'lxml')
        table = soup.find('div', class_='content__inner post__text')
        table = table.find_all('p', class_='text')

        flag = False
        flag_2 = False
        urls = []
        for resp in table:
            if resp.text.find('Заочная форма обучения') != -1:
                break
            if resp.text.find('Бакалавриат') != -1:
                #flag_2 = True
                pass
                #TODO: Очники
            if resp.text.find('Очно-заочная форма обучения') != -1:
                flag = True
            if flag_2:
                try:
                    for cont in resp.contents:
                        if str(cont).find('link') != -1 and cont.text.find('изм') != -1:
                            url = cont.attrs['href']
                            url = url[url.find('data') - 1:]
                            urls.append(self.BASE_URL + url)
                            flag_2 = False
                            break
                except:
                    pass
            if flag:
                try:
                    for cont in resp.contents:
                        if str(cont).find('link') != -1:
                            url = cont.attrs['href']
                            url = url[url.find('data') - 1:]
                            urls.append(self.BASE_URL + url)
                            break
                except:
                    pass
        return urls

    @staticmethod
    def get_html(url: str):
        session = requests.session()
        response = session.post(url)
        return response
